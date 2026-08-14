"""M3.10：Excel 模板导出/批量导入 + 知识文档导入。"""
import io
import zipfile

import pytest
from openpyxl import Workbook, load_workbook


@pytest.fixture(scope="module")
def ctx(client, admin_headers):
    owner = client.post("/api/members", json={"name": "导入负责人"}, headers=admin_headers).json()["data"]
    product_manager = client.post("/api/members", json={"name": "导入产品经理"}, headers=admin_headers).json()["data"]
    return {"member": owner["id"], "product_manager": product_manager["id"]}


def _fill(ws, headers_row3: list[list]):
    for row in headers_row3:
        ws.append(row)


def _xlsx(sheets: dict[str, list[list]]) -> bytes:
    """按模板结构造 xlsx：第1行表头占位、第2行说明占位、第3行起数据。"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(["h"] * 12)
        ws.append(["hint"] * 12)
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, headers, url, content, filename="import.xlsx"):
    return client.post(url, files={"file": (filename, content, "application/octet-stream")}, headers=headers)


def test_template_download(client, admin_headers):
    r = client.get("/api/itsm-import/vendor/template", headers=admin_headers)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "供应商" in wb.sheetnames
    assert wb["供应商"].cell(row=1, column=1).value == "*名称"
    # 枚举提示在说明行
    assert "A/B/C/D" in (wb["供应商"].cell(row=2, column=6).value or "")


def test_vendor_import_and_dedup(client, admin_headers, ctx):
    content = _xlsx({"供应商": [
        ["华云科技", "张三", "13800000000", "z@hy.com", "云资源", "A", ""],
        ["华云科技", "", "", "", "", "", ""],           # 同文件重复
        ["评级错误公司", "", "", "", "", "S", ""],       # 枚举非法
    ]})
    r = _upload(client, admin_headers, "/api/itsm-import/vendor", content).json()["data"]
    assert r["created"] == 1
    assert len(r["failed"]) == 2
    assert any("已存在" in e["error"] for e in r["failed"])
    assert any("可选值" in e["error"] for e in r["failed"])

    vendors = client.get("/api/vendors", headers=admin_headers).json()["data"]
    assert any(v["name"] == "华云科技" for v in vendors)


def test_contract_import_vendor_link(client, admin_headers, ctx):
    content = _xlsx({"合同": [
        ["云服务年度合同", "华云科技", 50, "2026-01-01", "2026-12-31", "导入负责人", ""],
        ["无供应商合同", "不存在的公司", 10, "2026-01-01", "2026-12-31", "", ""],
        ["日期反了", "华云科技", 10, "2026-12-31", "2026-01-01", "", ""],
    ]})
    r = _upload(client, admin_headers, "/api/itsm-import/contract", content).json()["data"]
    assert r["created"] == 1 and len(r["failed"]) == 2
    assert any("不存在" in e["error"] for e in r["failed"])
    assert any("结束日期早于开始日期" in e["error"] for e in r["failed"])


def test_catalog_two_sheet_import(client, admin_headers, ctx):
    content = _xlsx({
        "服务目录": [["基础设施服务", "gold", "机房与云", 1]],
        "服务项": [
            ["虚拟机申请", "基础设施服务", "日常运维", "导入负责人", "", 4, 24, "全员"],
            ["挂错目录的项", "没有这个目录", "", "", "", None, None, ""],
        ],
    })
    r = _upload(client, admin_headers, "/api/itsm-import/catalog", content).json()["data"]
    assert r["created"] == {"catalogs": 1, "items": 1}
    assert len(r["failed"]) == 1 and "不存在" in r["failed"][0]["error"]

    items = client.get("/api/service-items", headers=admin_headers).json()["data"]
    vm = next(i for i in items if i["name"] == "虚拟机申请")
    assert vm["catalog_name"] == "基础设施服务" and vm["owner_name"] == "导入负责人"


def test_ci_import(client, admin_headers, ctx):
    content = _xlsx({"配置项": [
        ["核心交换机", "网络", "导入负责人", "", "运行中", "生产", "", "华云科技", "", "2026-01-01", ""],
        ["负责人不存在的CI", "服务器", "查无此人", "", "", "", "", "", "", None, ""],
    ]})
    r = _upload(client, admin_headers, "/api/itsm-import/ci", content).json()["data"]
    assert r["created"] == 1 and len(r["failed"]) == 1
    cis = client.get("/api/cis", headers=admin_headers).json()["data"]
    sw = next(c for c in cis if c["name"] == "核心交换机")
    assert sw["category"] == "网络" and sw["vendor_name"] == "华云科技"


def test_ci_template_and_application_product_manager_import(client, admin_headers, ctx):
    template = client.get("/api/itsm-import/ci/template", headers=admin_headers)
    assert template.status_code == 200
    workbook = load_workbook(io.BytesIO(template.content))
    sheet = workbook["配置项"]
    assert sheet.cell(row=1, column=3).value == "*技术负责人姓名"
    assert sheet.cell(row=1, column=4).value == "应用产品经理姓名"
    hint = sheet.cell(row=2, column=4).value or ""
    assert "类别为应用" in hint and "必填" in hint

    content = _xlsx({"配置项": [
        ["导入应用产品经理", "应用", "导入负责人", "导入产品经理", "运行中", "生产", "", "", "", None, ""],
        ["缺少产品经理的应用", "应用", "导入负责人", "", "运行中", "生产", "", "", "", None, ""],
        ["未知产品经理的应用", "应用", "导入负责人", "查无此人", "运行中", "生产", "", "", "", None, ""],
        ["非应用填写产品经理", "网络", "导入负责人", "导入产品经理", "运行中", "生产", "", "", "", None, ""],
    ]})
    result = _upload(client, admin_headers, "/api/itsm-import/ci", content).json()["data"]
    assert result["created"] == 1 and len(result["failed"]) == 3
    errors = [row["error"] for row in result["failed"]]
    assert any("必须填写应用产品经理姓名" in error for error in errors)
    assert any("应用产品经理「查无此人」不是系统中的在岗人员" in error for error in errors)
    assert any("仅可填写在类别为应用" in error for error in errors)

    cis = client.get("/api/cis", headers=admin_headers).json()["data"]
    application = next(ci for ci in cis if ci["name"] == "导入应用产品经理")
    assert application["owner_name"] == "导入负责人"
    assert application["product_manager_id"] == ctx["product_manager"]
    assert application["product_manager_name"] == "导入产品经理"


def test_wrong_sheet_rejected(client, admin_headers):
    content = _xlsx({"随便什么表": [["x"]]})
    r = _upload(client, admin_headers, "/api/itsm-import/vendor", content).json()["data"]
    assert r["created"] == 0 and "缺少工作表" in r["failed"][0]["error"]


# ---------- 知识文档导入 ----------

def test_knowledge_md_import(client, admin_headers):
    md = "# 数据库备份手册\n\n## 每日备份\n\n执行 pg_dump。\n"
    r = client.post(
        "/api/knowledge/import",
        files={"file": ("backup.md", md.encode(), "text/markdown")},
        headers=admin_headers,
    ).json()["data"]
    assert r["title"] == "数据库备份手册"
    detail = client.get(f"/api/knowledge/{r['article_id']}", headers=admin_headers).json()["data"]
    assert detail["status"] == "draft" and detail["content_format"] == "markdown"
    assert "pg_dump" in detail["content"]


def _minimal_docx(paragraphs: list[str]) -> bytes:
    """手工构造最小 docx（避免引入 python-docx 依赖）。"""
    body = "".join(
        f'<w:p><w:pPr><w:pStyle w:val="{style}"/></w:pPr><w:r><w:t>{text}</w:t></w:r></w:p>'
        for style, text in paragraphs
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}</w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document)
    return buf.getvalue()


def test_knowledge_docx_import(client, admin_headers):
    docx = _minimal_docx([("Heading1", "运维值班制度"), ("Normal", "每周轮换，值班人负责 P1 响应。")])
    r = client.post(
        "/api/knowledge/import",
        files={"file": ("duty.docx", docx, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")},
        headers=admin_headers,
    ).json()["data"]
    assert r["title"] == "运维值班制度"
    detail = client.get(f"/api/knowledge/{r['article_id']}", headers=admin_headers).json()["data"]
    assert detail["content_format"] == "html"
    assert "<h1>" in detail["content"] and "值班人负责" in detail["content"]


def test_knowledge_html_sanitized(client, admin_headers):
    html = '<h1>安全文档</h1><script>alert(1)</script><p onclick="x">正文<img src="https://a/b.png"></p>'
    r = client.post(
        "/api/knowledge/import",
        files={"file": ("sec.html", html.encode(), "text/html")},
        headers=admin_headers,
    ).json()["data"]
    detail = client.get(f"/api/knowledge/{r['article_id']}", headers=admin_headers).json()["data"]
    assert "<script>" not in detail["content"] and "alert(1)" not in detail["content"]
    assert "onclick" not in detail["content"]
    assert "<img" in detail["content"]


def test_knowledge_unsupported_format(client, admin_headers):
    r = client.post(
        "/api/knowledge/import",
        files={"file": ("a.pdf", b"%PDF-1.4", "application/pdf")},
        headers=admin_headers,
    )
    assert r.json()["error"]["code"] == "IMPORT_FAILED"
