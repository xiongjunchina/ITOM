"""岗位定义与招聘需求 Excel 导入/导出闭环。"""

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


def _fill_xlsx(response, sheet_name: str, values: list):
    wb = load_workbook(BytesIO(response.content))
    ws = wb[sheet_name]
    for idx, value in enumerate(values, start=1):
        ws.cell(row=3, column=idx, value=value)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _with_invalid_column_dimension(content: bytes) -> bytes:
    """模拟在线表格导出器写入超大列号的工作簿。"""
    out = BytesIO()
    with ZipFile(BytesIO(content), "r") as source, ZipFile(out, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            raw = source.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                marker = b"<cols>"
                raw = raw.replace(marker, marker + b'<col min="63199" max="63199" width="14" customWidth="1"/>', 1)
            target.writestr(info, raw)
    return out.getvalue()


def test_position_template_import_and_export(client, admin_headers):
    template = client.get("/api/positions/template", headers=admin_headers)
    assert template.status_code == 200
    content = _fill_xlsx(
        template,
        "岗位定义",
        ["IT-OPS-001", "平台运维工程师", "运维", "基础架构；平台", "it_ops", "P4-P6", "广州", "Linux；K8s", 3, "否", "启用", 10, "负责平台稳定性与日常运维"],
    )
    imported = client.post(
        "/api/positions/import",
        headers=admin_headers,
        files={"file": ("positions.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["data"]["created"] == 1
    row = client.get("/api/positions?q=IT-OPS-001", headers=admin_headers).json()["data"][0]
    assert row["position_family"] == "运维"
    assert row["service_domains"] == ["基础架构", "平台"]
    assert row["headcount"] == 3

    exported = client.get("/api/positions/export", headers=admin_headers)
    assert exported.status_code == 200
    assert "岗位定义" in load_workbook(BytesIO(exported.content)).sheetnames


def test_hiring_template_import_and_export(client, admin_headers):
    position = client.post(
        "/api/positions", json={"position_code": "IT-DATA-001", "name": "数据工程师", "headcount": 2}, headers=admin_headers
    ).json()["data"]
    template = client.get("/api/hiring-needs/template", headers=admin_headers)
    content = _fill_xlsx(
        template,
        "招聘需求",
        [None, "IT-DATA-001", position["name"], "中级", 1, "熟悉数据建模与质量治理", "待招聘", "首批招聘"],
    )
    imported = client.post(
        "/api/hiring-needs/import",
        headers=admin_headers,
        files={"file": ("hiring.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["data"]["created"] == 1
    rows = client.get("/api/hiring-needs", headers=admin_headers).json()["data"]
    assert any(row["position_code"] == "IT-DATA-001" for row in rows)
    exported = client.get("/api/hiring-needs/export", headers=admin_headers)
    assert exported.status_code == 200
    assert "招聘需求" in load_workbook(BytesIO(exported.content)).sheetnames


def test_exported_position_can_be_edited_and_reimported(client, admin_headers):
    """导出的回导模板应按编码更新，而不是因为编码/名称匹配冲突返回 500。"""
    created = client.post(
        "/api/positions",
        json={"position_code": "IT-REIMPORT-001", "name": "回导岗位", "headcount": 1},
        headers=admin_headers,
    )
    assert created.status_code == 200, created.text
    position_id = created.json()["data"]["id"]

    exported = client.get("/api/positions/export", headers=admin_headers)
    wb = load_workbook(BytesIO(exported.content))
    ws = wb["岗位定义"]
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "IT-REIMPORT-001":
            ws.cell(row=row, column=5, value="platform_ops")
            ws.cell(row=row, column=9, value=4)
            break
    out = BytesIO()
    wb.save(out)

    imported = client.post(
        "/api/positions/import",
        headers=admin_headers,
        files={"file": ("positions-export.xlsx", out.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["data"]["updated"] >= 1
    row = client.get(f"/api/positions?q=IT-REIMPORT-001", headers=admin_headers).json()["data"][0]
    assert row["id"] == position_id
    assert row["primary_roles"] == ["platform_ops"]
    assert row["headcount"] == 4


def test_import_rejects_invalid_workbook_without_500(client, admin_headers):
    for endpoint, filename in (("/api/positions/import", "positions.xlsx"), ("/api/hiring-needs/import", "hiring.xlsx")):
        response = client.post(endpoint, headers=admin_headers, files={"file": (filename, b"not-an-xlsx", "application/octet-stream")})
        assert response.status_code == 200, response.text
        data = response.json()["data"]
        assert data["created"] == 0
        assert data["failed"][0]["row"] == 0
        assert "有效的 Excel" in data["failed"][0]["error"]


def test_import_repairs_invalid_column_dimension(client, admin_headers):
    template = client.get("/api/positions/template", headers=admin_headers)
    content = _fill_xlsx(
        template,
        "岗位定义",
        ["IT-BROKEN-XLSX", "非法列宽测试岗位", "运维", "基础设施", "it_ops", "P4", "广州", "Linux", 1, "否", "启用", 1, "测试可恢复的工作簿"],
    )
    content = _with_invalid_column_dimension(content)
    imported = client.post(
        "/api/positions/import",
        headers=admin_headers,
        files={"file": ("positions-broken-columns.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["data"]["created"] == 1


def test_cio_can_inline_update_and_delete_position_and_hiring(client, admin_headers):
    user = client.post(
        "/api/admin/users",
        json={"username": "position_cio", "password": "pass123", "roles": ["cio"]},
        headers=admin_headers,
    )
    assert user.status_code == 200, user.text
    token = client.post("/api/auth/login", json={"username": "position_cio", "password": "pass123"}).json()["data"]["token"]
    cio_headers = {"Authorization": f"Bearer {token}"}

    position = client.post("/api/positions", json={"name": "可编辑岗位", "headcount": 1}, headers=cio_headers).json()["data"]
    updated = client.patch(f"/api/positions/{position['id']}", json={"name": "已更新岗位"}, headers=cio_headers)
    assert updated.status_code == 200, updated.text
    hiring = client.post(
        "/api/hiring-needs",
        json={"position_id": position["id"], "headcount": 1, "qualification": "熟悉 IT 服务管理"},
        headers=cio_headers,
    ).json()["data"]
    assert client.delete(f"/api/positions/{position['id']}", headers=cio_headers).status_code == 400
    assert client.delete(f"/api/hiring-needs/{hiring['id']}", headers=cio_headers).status_code == 200
    assert client.delete(f"/api/positions/{position['id']}", headers=cio_headers).status_code == 200
