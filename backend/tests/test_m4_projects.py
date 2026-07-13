"""M4：项目/组合/WBS/里程碑/风险/成本/章程导入/派生指标。"""
import io
import zipfile
from datetime import date, timedelta

import pytest

TODAY = date.today()


@pytest.fixture(scope="module")
def ctx(client, admin_headers):
    def member_and_user(name, username, roles):
        m = client.post("/api/members", json={"name": name}, headers=admin_headers).json()["data"]
        client.post(
            "/api/admin/users",
            json={"username": username, "password": "pass123", "roles": roles, "person_id": m["id"]},
            headers=admin_headers,
        )
        token = client.post("/api/auth/login", json={"username": username, "password": "pass123"}).json()["data"]["token"]
        return m["id"], {"Authorization": f"Bearer {token}"}

    pm_person, pm = member_and_user("项目张经理", "pm01", ["it_pm"])
    dev_person, dev = member_and_user("开发小李", "dev01", ["it_dev"])
    cio_person, cio = member_and_user("陈CIO", "cio01", ["cio"])
    return {"pm_person": pm_person, "pm": pm, "dev_person": dev_person, "dev": dev,
            "cio_person": cio_person, "cio": cio, "member_and_user": member_and_user}


def _mkproject(client, headers, **kw):
    payload = {
        "name": "测试项目A", "pm": kw.pop("pm"),
        "planned_start": str(TODAY - timedelta(days=10)),
        "planned_end": str(TODAY + timedelta(days=20)),
        **kw,
    }
    r = client.post("/api/projects", json=payload, headers=headers)
    assert r.json()["success"], r.text
    return r.json()["data"]


def test_project_create_and_flow(client, ctx):
    p = _mkproject(client, ctx["pm"], pm=ctx["pm_person"])
    assert p["project_code"].startswith("PJ-") and p["status"] == "planning"

    detail = client.get(f"/api/projects/{p['id']}", headers=ctx["pm"]).json()["data"]
    assert detail["process"]["definition_name"] == "项目关键节点流程"
    assert detail["can_edit"] is True

    # it_dev 无 projects.create
    r = client.post("/api/projects", json={"name": "越权项目", "pm": ctx["dev_person"],
                                           "planned_start": str(TODAY), "planned_end": str(TODAY)}, headers=ctx["dev"])
    assert r.status_code == 403

    # 流转：规划中→进行中 打点 actual_start
    r = client.post(f"/api/projects/{p['id']}/transition", json={"to": "active", "fields": {}}, headers=ctx["pm"])
    assert r.json()["data"]["status"] == "active"
    detail = client.get(f"/api/projects/{p['id']}", headers=ctx["pm"]).json()["data"]
    assert detail["actual_start"] is not None


def test_wbs_progress_and_health(client, ctx):
    p = _mkproject(client, ctx["pm"], pm=ctx["pm_person"], name="进度项目", budget_10k=100)
    pid = p["id"]
    client.post(f"/api/projects/{pid}/transition", json={"to": "active", "fields": {}}, headers=ctx["pm"])

    # 两个等工期任务：一个已完成（过去），一个未开始（未来）→ 进度 50%
    t1 = client.post(f"/api/projects/{pid}/wbs", json={
        "name": "已完成任务", "assignee": ctx["dev_person"],
        "start_date": str(TODAY - timedelta(days=9)), "end_date": str(TODAY - timedelta(days=5)),
    }, headers=ctx["pm"]).json()["data"]
    client.post(f"/api/projects/{pid}/wbs", json={
        "name": "未来任务", "assignee": ctx["dev_person"],
        "start_date": str(TODAY + timedelta(days=5)), "end_date": str(TODAY + timedelta(days=9)),
    }, headers=ctx["pm"]).json()["data"]

    # 任务负责人自己更新完成度（数据范围规则）
    r = client.patch(f"/api/wbs/{t1['id']}", json={"progress": 100}, headers=ctx["dev"])
    assert r.json()["success"], r.text
    # 负责人改其他字段被拒
    r = client.patch(f"/api/wbs/{t1['id']}", json={"name": "改名"}, headers=ctx["dev"])
    assert r.status_code == 403

    detail = client.get(f"/api/projects/{pid}", headers=ctx["pm"]).json()["data"]
    assert detail["progress"] == 50.0
    assert detail["health"] == "green"  # 计划=实际（过去任务完成，未来任务未到期）
    assert detail["task_done"] == 1

    # WBS 层级编码
    child = client.post(f"/api/projects/{pid}/wbs", json={
        "name": "子任务", "assignee": ctx["dev_person"], "parent_task_id": t1["id"],
        "start_date": str(TODAY), "end_date": str(TODAY + timedelta(days=1)),
    }, headers=ctx["pm"]).json()["data"]
    wbs = client.get(f"/api/projects/{pid}/wbs", headers=ctx["pm"]).json()["data"]
    codes = {w["name"]: w["wbs_code"] for w in wbs}
    assert codes["已完成任务"] == "1" and codes["子任务"] == "1.1" and codes["未来任务"] == "2"


def test_health_yellow_on_overdue_milestone_and_red_risk(client, ctx):
    p = _mkproject(client, ctx["pm"], pm=ctx["pm_person"], name="健康度项目")
    pid = p["id"]
    client.post(f"/api/projects/{pid}/transition", json={"to": "active", "fields": {}}, headers=ctx["pm"])

    # 一个按期完成的大任务（占大工期，稀释整体偏差到黄色区间）
    big = client.post(f"/api/projects/{pid}/wbs", json={
        "name": "主体工作按期完成", "assignee": ctx["dev_person"],
        "start_date": str(TODAY - timedelta(days=30)), "end_date": str(TODAY - timedelta(days=1)),
    }, headers=ctx["pm"]).json()["data"]
    client.patch(f"/api/wbs/{big['id']}", json={"progress": 100}, headers=ctx["pm"])
    # 里程碑=WBS 勾选「是」；小任务、计划结束已过且未完成 → 已延期 → 逾期里程碑（触发黄）
    ms_task = client.post(f"/api/projects/{pid}/wbs", json={
        "name": "已逾期里程碑", "assignee": ctx["dev_person"], "is_milestone": True,
        "start_date": str(TODAY - timedelta(days=1)), "end_date": str(TODAY - timedelta(days=1)),
    }, headers=ctx["pm"]).json()["data"]
    detail = client.get(f"/api/projects/{pid}", headers=ctx["pm"]).json()["data"]
    assert detail["health"] == "yellow" and detail["milestone_overdue"] == 1
    # 里程碑跟踪派生视图汇总该行
    track = client.get(f"/api/projects/{pid}/milestone-tracking", headers=ctx["pm"]).json()["data"]
    assert len(track) == 1 and track[0]["name"] == "已逾期里程碑" and track[0]["status"] == "已延期"

    client.post(f"/api/projects/{pid}/risks", json={
        "title": "核心供应商跑路", "probability": "高", "impact": "高", "mitigation": "备选供应商",
    }, headers=ctx["pm"])
    detail = client.get(f"/api/projects/{pid}", headers=ctx["pm"]).json()["data"]
    assert detail["health"] == "red" and detail["red_risks"] == 1

    # 关闭风险 + 里程碑完成度置 100 → 恢复绿
    risks = client.get(f"/api/projects/{pid}/risks", headers=ctx["pm"]).json()["data"]
    client.patch(f"/api/risks/{risks[0]['id']}", json={"status": "已关闭"}, headers=ctx["pm"])
    client.patch(f"/api/wbs/{ms_task['id']}", json={"progress": 100}, headers=ctx["pm"])
    detail = client.get(f"/api/projects/{pid}", headers=ctx["pm"]).json()["data"]
    assert detail["health"] == "green"


def test_cost_and_budget_usage(client, ctx):
    p = _mkproject(client, ctx["pm"], pm=ctx["pm_person"], name="成本项目", budget_10k=50)
    pid = p["id"]
    client.post(f"/api/projects/{pid}/costs", json={"entry_date": str(TODAY), "amount_10k": 10, "note": "外包"}, headers=ctx["pm"])
    client.post(f"/api/projects/{pid}/costs", json={"entry_date": str(TODAY), "amount_10k": 15}, headers=ctx["pm"])
    detail = client.get(f"/api/projects/{pid}", headers=ctx["pm"]).json()["data"]
    assert detail["actual_cost_10k"] == 25 and detail["budget_usage"] == 50.0


def test_portfolio(client, ctx):
    r = client.post("/api/portfolios", json={"name": "数字化转型", "owner_id": ctx["cio_person"], "year": "2026"}, headers=ctx["cio"])
    assert r.json()["success"], r.text
    pf = r.json()["data"]["id"]
    _mkproject(client, ctx["pm"], pm=ctx["pm_person"], name="组合内项目", portfolio_id=pf)
    rows = client.get("/api/portfolios", headers=ctx["pm"]).json()["data"]
    row = next(x for x in rows if x["name"] == "数字化转型")
    assert row["project_count"] == 1 and row["owner_name"] == "陈CIO"


def _charter_docx() -> bytes:
    """构造最小章程 docx：字段表 + WBS 表 + 风险节。"""
    def tbl(rows):
        out = "<w:tbl>"
        for cells in rows:
            out += "<w:tr>" + "".join(
                f"<w:tc><w:p><w:r><w:t>{c}</w:t></w:r></w:p></w:tc>" for c in cells
            ) + "</w:tr>"
        return out + "</w:tbl>"

    def p(text):
        return f"<w:p><w:r><w:t>{text}</w:t></w:r></w:p>"

    start = (TODAY + timedelta(days=1)).isoformat()
    end = (TODAY + timedelta(days=60)).isoformat()
    e1 = (TODAY + timedelta(days=10)).isoformat()
    e2 = (TODAY + timedelta(days=40)).isoformat()
    body = (
        tbl([["项目名称", "数据中台一期"], ["项目经理", "项目张经理"],
             ["计划开始", start], ["计划完成", end], ["项目预算", "120万元"]])
        + p("1. 项目背景") + p("统一数据口径，减少重复建设。")
        + p("3. 项目目标") + p("上线统一数仓与指标服务。")
        + p("5. WBS 任务分解与里程碑")
        # 10 列：阶段|WBS编号|任务名称|词典说明|DoD|责任人|里程碑|前置(WBS号)|计划开始|计划结束
        + tbl([["阶段", "WBS编号", "任务名称(交付物)", "WBS词典说明", "交付物/DoD", "责任人", "里程碑", "前置任务(WBS号)", "计划开始", "计划结束"],
               ["1.需求", "1", "需求调研", "访谈业务部门", "调研报告", "", "否", "", start, e1],
               ["2.建设", "2", "平台搭建", "部署数仓", "可用环境（里程碑）", "", "是", "1", (TODAY + timedelta(days=11)).isoformat(), e2]])
        + p("7.1 关键风险")
        + tbl([["风险类别", "风险描述", "概率", "影响", "应对措施"],  # 表头
               ["技术风险", "数据源接入复杂", "高", "中", "预留缓冲期"],
               ["资源风险", "关键人员不足", "中", "高", "提前锁定资源"]])
        + p("7.2 关键假设")
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}</w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/></Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document)
    return buf.getvalue()


def test_charter_parse_and_create(client, ctx):
    r = client.post(
        "/api/projects/charter/parse",
        files={"file": ("charter.docx", _charter_docx())},
        headers=ctx["pm"],
    ).json()["data"]
    f = r["fields"]
    assert f["name"] == "数据中台一期" and f["budget_10k"] == 120
    assert f["pm"] == ctx["pm_person"]  # 姓名解析到人员
    assert len(r["drafts"]["wbs"]) == 2 and len(r["drafts"]["risks"]) == 2  # 里程碑=WBS 派生，无独立 milestones
    assert "项目背景" in f["description"]
    assert sum(1 for w in r["drafts"]["wbs"] if w["is_milestone"]) == 1  # 平台搭建 里程碑=是

    r2 = client.post("/api/projects/charter/create", json={
        "fields": f, "wbs": r["drafts"]["wbs"], "risks": r["drafts"]["risks"],
    }, headers=ctx["pm"]).json()["data"]
    assert r2["created"] == {"wbs": 2, "milestones": 1, "risks": 2}  # milestones = 里程碑=是 的行数

    detail = client.get(f"/api/projects/{r2['project_id']}", headers=ctx["pm"]).json()["data"]
    assert detail["name"] == "数据中台一期" and detail["task_total"] == 2
    wbs = client.get(f"/api/projects/{r2['project_id']}/wbs", headers=ctx["pm"]).json()["data"]
    assert [w["wbs_code"] for w in wbs] == ["1", "2"]
    # 里程碑跟踪派生：仅 平台搭建（里程碑=是）
    track = client.get(f"/api/projects/{r2['project_id']}/milestone-tracking", headers=ctx["pm"]).json()["data"]
    assert len(track) == 1 and track[0]["name"] == "平台搭建"


def test_dashboard_project_section(client, ctx):
    dash = client.get("/api/dashboard", headers=ctx["pm"]).json()["data"]
    assert dash["project"]["active"] >= 2
    assert sum(dash["project"]["health"].values()) >= 3


def test_progress_template_and_import(client, ctx):
    """M9：进度导入模板单 WBS 表（WBS编号层级+前置、里程碑=WBS 标志、完成度%）。"""
    import io
    from openpyxl import Workbook, load_workbook

    r = client.get("/api/project-progress/template", headers=ctx["pm"])
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"WBS任务"}  # 里程碑=WBS 派生，不再单独出表

    p = _mkproject(client, ctx["pm"], pm=ctx["pm_person"], name="导入进度项目")
    pid = p["id"]

    wb2 = Workbook(); wb2.remove(wb2.active)
    ws = wb2.create_sheet("WBS任务")
    # 14 列：阶段|WBS编号|任务名称|词典说明|DoD|责任人|里程碑|前置(WBS号)|计划开始|计划结束|实际开始|实际结束|完成度%|备注
    ws.append(["h"] * 14); ws.append(["hint"] * 14)
    ws.append(["1.需求", "1", "需求调研", "访谈", "调研报告", "项目张经理", "否", "", str(TODAY), str(TODAY + timedelta(days=5)), "", "", "100", ""])
    ws.append(["2.建设", "2", "方案设计", "原型", "设计文档", "项目张经理", "否", "1", str(TODAY + timedelta(days=6)), str(TODAY + timedelta(days=10)), "", "", "50", ""])
    ws.append(["2.建设", "2.1", "接口开发", "接口", "接口代码", "项目张经理", "否", "", str(TODAY + timedelta(days=6)), str(TODAY + timedelta(days=15)), "", "", "", ""])
    ws.append(["3.上线", "3", "一期上线", "上线", "上线报告", "项目张经理", "是", "2", str(TODAY + timedelta(days=16)), str(TODAY + timedelta(days=20)), "", "", "", ""])
    ws.append(["", "", "坏行", "", "", "查无此人", "否", "", str(TODAY), str(TODAY), "", "", "", ""])
    buf = io.BytesIO(); wb2.save(buf)

    result = client.post(
        f"/api/projects/{pid}/import-progress",
        files={"file": ("progress.xlsx", buf.getvalue())},
        headers=ctx["pm"],
    ).json()["data"]
    assert result["created"] == {"wbs": 4, "milestones": 1}  # 一期上线 里程碑=是
    assert any("查无此人" in e["error"] for e in result["failed"])

    wbs = client.get(f"/api/projects/{pid}/wbs", headers=ctx["pm"]).json()["data"]
    by_name = {w["name"]: w for w in wbs}
    assert by_name["接口开发"]["parent_task_id"] == by_name["方案设计"]["id"]  # 层级由 WBS编号 建立
    assert by_name["方案设计"]["predecessor_ids"] == [by_name["需求调研"]["id"]]  # 前置按编号
    assert "." in by_name["接口开发"]["wbs_code"]  # 子任务层级编码
    assert by_name["需求调研"]["progress"] == 100 and by_name["方案设计"]["progress"] == 50
    # 里程碑跟踪派生
    track = client.get(f"/api/projects/{pid}/milestone-tracking", headers=ctx["pm"]).json()["data"]
    assert [t["name"] for t in track] == ["一期上线"]

    # 示例项目禁止导入
    projects = client.get("/api/projects", headers=ctx["pm"]).json()["data"]
    demo = next(x for x in projects if x.get("is_example"))
    r = client.post(f"/api/projects/{demo['id']}/import-progress",
                    files={"file": ("p.xlsx", buf.getvalue())}, headers=ctx["pm"])
    assert r.json()["error"]["code"] == "EXAMPLE_READONLY"
