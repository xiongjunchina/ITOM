"""M10：需求六维加权评分 + 四象限 + 评估门 + 评分配置 + 批量导入。"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from app.db import SessionLocal
from app.models import Requirement


@pytest.fixture(scope="module")
def ctx(client, admin_headers):
    def member_and_user(name, username, roles):
        m = client.post("/api/members", json={"name": name}, headers=admin_headers).json()["data"]
        client.post("/api/admin/users",
                    json={"username": username, "password": "pass123", "roles": roles, "person_id": m["id"]},
                    headers=admin_headers)
        token = client.post("/api/auth/login", json={"username": username, "password": "pass123"}).json()["data"]["token"]
        return m["id"], {"Authorization": f"Bearer {token}"}

    pdm_p, pdm = member_and_user("产品M10", "pdm_m10", ["it_pdm"])
    domain = client.post("/api/admin/business-domains",
                         json={"code": "dig10", "name": "数字化业务线", "owner_id": None},
                         headers=admin_headers).json()["data"]
    return {"pdm": pdm, "pdm_p": pdm_p, "admin": admin_headers, "domain": domain["id"]}


def _register(client, headers, domain, **kw):
    payload = {"title": "多平台ERP中台", "req_type": "功能", "business_domain_id": domain,
               "description": "支撑多平台运营的中台", **kw}
    return client.post("/api/requirements", json=payload, headers=headers).json()["data"]


def test_score_weighted_total_and_quadrant(client, ctx):
    r = _register(client, ctx["pdm"], ctx["domain"])
    rid = r["id"]
    client.post(f"/api/requirements/{rid}/transition", json={"to": "evaluating", "fields": {}}, headers=ctx["admin"])

    # 六维评分 → 加权总分 3.8，象限=战略下注
    resp = client.post(f"/api/requirements/{rid}/score", json={
        "d1_strategy": 5, "d2_value": 5, "d3_tech": 3, "d4_org": 3, "d5_risk": 3, "d6_speed": 3,
        "decision": "通过", "comment": "战略基建",
    }, headers=ctx["admin"])
    data = resp.json()["data"]
    assert data["weighted_total"] == 3.8, resp.text
    assert data["quadrant"] == "战略下注"
    assert data["decision"] == "通过"

    # 列表/详情回填
    detail = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    assert detail["d1_strategy"] == 5 and detail["weighted_total"] == 3.8
    assert len(detail["scores"]) == 1 and detail["scores"][0]["is_consensus"] is True


def test_detail_exposes_persisted_scores_without_history_and_rejects_with_reason(client, ctx):
    """历史/导入需求可回填评分；首审批节点驳回后由登记人补充。"""
    r = _register(client, ctx["pdm"], ctx["domain"], title="历史主表评分需求")
    rid = r["id"]
    with SessionLocal() as db:
        requirement = db.get(Requirement, rid)
        requirement.score_d1_strategy = 5
        requirement.score_d2_value = 4
        requirement.score_d3_tech = 4
        requirement.score_d4_org = 3
        requirement.score_d5_risk = 2
        requirement.score_d6_speed = 4
        db.commit()

    detail = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    assert detail["scores"] == []
    assert [detail[key] for key in ("d1_strategy", "d2_value", "d3_tech", "d4_org", "d5_risk", "d6_speed")] == [5, 4, 4, 3, 2, 4]
    assert detail["weighted_total"] == 4.1 and detail["quadrant"] == "战略下注"

    rejected = client.post(
        f"/api/requirements/{rid}/score",
        json={"decision": "驳回", "comment": "当前业务优先级已调整，停止投入实施"},
        headers=ctx["admin"],
    )
    assert rejected.status_code == 200, rejected.text
    assert rejected.json()["data"]["status"] == "supplementing"
    returned = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    assert returned["process"]["status"] == "returned"
    assert returned["can_resubmit"] is True
    assert returned["process"]["return_info"]["reason"] == "当前业务优先级已调整，停止投入实施"
    assert all(step["task_status"] != "待处理" for step in returned["process"]["steps"])


def test_eval_gate_quadrant_and_auto_flow(client, ctx):
    """M16 评估门：未评分不可立项；重评象限仅 搁置/驳回（驳回必填理由）；决议即自动流转。"""
    r = _register(client, ctx["pdm"], ctx["domain"], title="低优先需求")
    rid = r["id"]
    assert r["status"] == "evaluating"  # M16：登记即进评审

    # 未评分就立项 → 拦截
    resp = client.post(f"/api/requirements/{rid}/score", json={"decision": "通过"}, headers=ctx["admin"])
    assert resp.json()["error"]["code"] == "EVAL_INCOMPLETE"

    # 低分落「重新评估」象限 → 立项被象限约束拦截
    resp = client.post(f"/api/requirements/{rid}/score", json={
        "d1_strategy": 2, "d2_value": 2, "d3_tech": 3, "d4_org": 3, "d5_risk": 3, "d6_speed": 2,
        "decision": "通过",
    }, headers=ctx["admin"])
    assert resp.json()["error"]["code"] == "QUADRANT_REJECTED"

    # 驳回必填理由
    resp = client.post(f"/api/requirements/{rid}/score", json={"decision": "驳回"}, headers=ctx["admin"])
    assert resp.json()["error"]["code"] == "REASON_REQUIRED"

    # 搁置 → 自动流转 on_hold（补充后可重新评审）
    resp = client.post(f"/api/requirements/{rid}/score", json={"decision": "搁置", "comment": "价值论证不足，请补充预期收益"}, headers=ctx["admin"])
    assert resp.json()["data"]["status"] == "on_hold"

    # 重新进入评审 → 提分至非重评象限 → 立项自动流转 analyzing
    client.post(f"/api/requirements/{rid}/transition", json={"to": "evaluating", "fields": {}}, headers=ctx["admin"])
    resp = client.post(f"/api/requirements/{rid}/score", json={
        "d1_strategy": 4, "d2_value": 4, "d3_tech": 4, "d4_org": 3, "d5_risk": 2, "d6_speed": 4,
        "decision": "通过",
    }, headers=ctx["admin"])
    data = resp.json()["data"]
    assert data["status"] == "analyzing" and data["flowed_to"] == "analyzing"


def test_reject_returns_to_requester_and_resubmits_same_instance(client, ctx):
    """M108：首审批节点驳回→登记人补充→原需求、原流程实例重新提交。"""
    r = _register(client, ctx["pdm"], ctx["domain"], title="被驳回需求")
    rid = r["id"]
    before = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    instance_id = before["process"]["id"]
    resp = client.post(f"/api/requirements/{rid}/score", json={
        "d1_strategy": 1, "d2_value": 2, "d3_tech": 3, "d4_org": 3, "d5_risk": 4, "d6_speed": 2,
        "decision": "驳回", "comment": "与年度战略无关且价值不可量化",
    }, headers=ctx["admin"])
    assert resp.json()["data"]["status"] == "supplementing", resp.text
    detail = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    assert detail["closure_note"] is None
    assert detail["decision"] is None
    assert detail["weighted_total"] is None
    assert detail["process"]["id"] == instance_id

    patched = client.patch(
        f"/api/requirements/{rid}",
        json={"description": "已补充业务价值测算和附件说明"},
        headers=ctx["pdm"],
    )
    assert patched.status_code == 200, patched.text
    resubmitted = client.post(f"/api/requirements/{rid}/resubmit", json={}, headers=ctx["pdm"])
    assert resubmitted.status_code == 200, resubmitted.text
    assert resubmitted.json()["data"]["instance_id"] == instance_id
    after = client.get(f"/api/requirements/{rid}", headers=ctx["admin"]).json()["data"]
    assert after["status"] == "evaluating"
    assert after["process"]["status"] == "running"
    assert after["process"]["current_step_seq"] == 1
    current = next(step for step in after["process"]["steps"] if step["seq"] == 1)
    assert current["task_status"] == "待处理"


def test_scoring_config_admin_only(client, ctx):
    cfg = client.get("/api/requirements/scoring-config", headers=ctx["admin"]).json()["data"]
    assert cfg["weights"]["d1"] == 0.2 and cfg["thresholds"]["total"] == 3.5

    # 非管理员不可改
    resp = client.put("/api/requirements/scoring-config", json={"thresholds": {"total": 4.0, "strategic": 4, "viable": 3}}, headers=ctx["pdm"])
    assert resp.json()["error"]["code"] == "FORBIDDEN"

    # 权重和不为 1 → 拒绝
    resp = client.put("/api/requirements/scoring-config",
                      json={"weights": {"d1": 0.5, "d2": 0.2, "d3": 0.2, "d4": 0.1, "d5": 0.1, "d6": 0.2}},
                      headers=ctx["admin"])
    assert resp.json()["error"]["code"] == "INVALID_WEIGHTS"

    # admin 改阈值成功
    resp = client.put("/api/requirements/scoring-config",
                      json={"thresholds": {"total": 3.8, "strategic": 4, "viable": 3}}, headers=ctx["admin"])
    assert resp.json()["success"]


def test_template_download_and_import(client, ctx):
    # 下载模板
    tpl = client.get("/api/requirements/template", headers=ctx["admin"])
    assert tpl.status_code == 200 and "spreadsheetml" in tpl.headers["content-type"]

    # 新模板只包含登记阶段可填写的 8 列，不再把渠道部门、六维评分、决议和人天带给登记人。
    wb = load_workbook(io.BytesIO(tpl.content))
    ws = wb["需求登记"]
    assert ws.max_column == 8
    assert all("战略对齐" not in str(cell.value) for cell in ws[1])
    ws.append(["海外仓WMS登记", "功能", "数字化业务线", "库存效率中台", None, None, "补货精度提升", "库存周转"])
    ws.append(["达人库", "业务", "数字化业务线", "达人资源库", None, None, None, None])
    ws.append(["坏行", "功能", "不存在的业务线", "描述", None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post("/api/requirements/import",
                       files={"file": ("req.xlsx", buf.getvalue(),
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                       headers=ctx["admin"])
    data = resp.json()["data"]
    assert data["imported"] == 2, resp.text
    assert len(data["errors"]) == 1 and "不存在" in data["errors"][0]["error"]

    # 新版登记模板落到评估中，但评分和决议由后续流程节点填写。
    listing = client.get("/api/requirements?q=海外仓", headers=ctx["admin"]).json()["data"]
    registration = next(x for x in listing if x["title"] == "海外仓WMS登记")
    assert registration["status"] == "evaluating" and registration["decision"] is None

    # 旧版模板仍可导入，避免历史下载文件因模板收敛而失效。
    legacy = Workbook()
    legacy.remove(legacy.active)
    legacy_ws = legacy.create_sheet("需求登记")
    legacy_headers = [
        "*需求名称", "*需求类型", "*所属业务域", "*需求描述", "需求来源", "渠道/部门", "提出人",
        "期望完成时间", "期望效果", "运营价值", "战略对齐(1-5)", "业务价值(1-5)", "技术可行性(1-5)",
        "组织就绪(1-5)", "风险等级(1-5)", "价值速度(1-5)", "最终决议", "PRD人天", "开发人天",
    ]
    legacy_ws.append(legacy_headers)
    legacy_ws.append([None] * len(legacy_headers))
    legacy_ws.append(["海外仓WMS", "功能", "数字化业务线", "历史导入", None, "数字化", "李强",
                      None, "补货精度提升", "库存周转", 5, 5, 4, 3, 2, 4, "通过", 20, 60])
    legacy_buf = io.BytesIO()
    legacy.save(legacy_buf)
    legacy_resp = client.post(
        "/api/requirements/import",
        files={"file": ("legacy.xlsx", legacy_buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ctx["admin"],
    )
    assert legacy_resp.json()["data"]["imported"] == 1, legacy_resp.text
    listing = client.get("/api/requirements?q=海外仓", headers=ctx["admin"]).json()["data"]
    wms = next(x for x in listing if x["title"] == "海外仓WMS")
    assert wms["status"] == "evaluating" and wms["decision"] == "通过"
    # 5,5,4,3,2,4 → 0.2*5+0.2*5+0.2*4+0.1*3+0.1*(6-2)+0.2*4 = 1+1+0.8+0.3+0.4+0.8 = 4.3
    assert wms["weighted_total"] == 4.3 and wms["quadrant"] == "战略下注"


def test_active_tasks_board(client, ctx):
    """需求走到实现阶段，任务(含描述/工天)出现在实现任务清单。"""
    r = _register(client, ctx["pdm"], ctx["domain"], title="任务看板需求")
    rid = r["id"]
    client.post(f"/api/requirements/{rid}/transition", json={"to": "evaluating", "fields": {}}, headers=ctx["admin"])
    client.post(f"/api/requirements/{rid}/score", json={
        "d1_strategy": 5, "d2_value": 5, "d3_tech": 4, "d4_org": 4, "d5_risk": 2, "d6_speed": 5, "decision": "通过",
    }, headers=ctx["admin"])
    client.post(f"/api/requirements/{rid}/transition", json={"to": "analyzing", "fields": {}}, headers=ctx["admin"])
    client.patch(f"/api/requirements/{rid}", json={"owner": ctx["pdm_p"], "solution": "MVP"}, headers=ctx["admin"])
    resp = client.post(f"/api/requirements/{rid}/transition", json={"to": "implementing", "fields": {}}, headers=ctx["admin"])
    assert resp.json()["data"]["status"] == "implementing", resp.text

    client.post(f"/api/requirements/{rid}/tasks", json={
        "name": "接口开发", "description": "对接海外仓API", "assignee": ctx["pdm_p"], "plan_effort": 5,
    }, headers=ctx["admin"])

    rows = client.get("/api/requirements/tasks/active", headers=ctx["admin"]).json()["data"]
    mine = next(x for x in rows if x["requirement_id"] == rid)
    assert mine["description"] == "对接海外仓API" and mine["plan_effort"] == 5
    assert mine["requirement_code"] == r["requirement_code"] and mine["requirement_status"] == "implementing"
    assert mine["name"] == "接口开发"
