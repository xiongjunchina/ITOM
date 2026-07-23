"""Excel 模板导出与批量导入通用引擎（M3.10）。

- build_template(): 生成空白 xlsx——首行表头（必填加 *），第二行灰色说明（含枚举提示），列宽自适应
- parse_sheet(): 逐行读取校验——必填/枚举/日期/数字/自定义解析器；返回 (有效行, 错误列表[行号+原因])
- 导入语义：有效行入库、失败行报告（部分成功），重复判定由各实体导入器负责
"""
import io
from zipfile import BadZipFile
from zipfile import ZIP_DEFLATED, ZipFile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Callable
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class Col:
    key: str
    label: str
    required: bool = False
    hint: str = ""
    enum: list[str] | None = None
    kind: str = "str"  # str/float/int/date
    max_length: int | None = None


@dataclass
class Sheet:
    name: str
    columns: list[Col] = field(default_factory=list)


_XLSX_MAX_COLUMN = 18278  # openpyxl.utils.cell.get_column_letter 的最大列号（ZZZ）
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _repair_invalid_column_dimensions(file_bytes: bytes) -> bytes | None:
    """删除部分表格软件生成的非法 ``<col min=...>`` 节点。

    某些在线表格导出器会把列宽的 min/max 写成超出 Excel 上限的值（例如
    63199）。这类文件在 Excel 中仍能打开，但 openpyxl 会在读取列定义时
    直接抛 ValueError。删除异常列宽只会丢失该列的自定义宽度，不影响单元格
    数据本身，之后由导入字段定义完成校验。
    """
    try:
        with ZipFile(io.BytesIO(file_bytes), "r") as source:
            names = source.namelist()
            sheet_names = [n for n in names if n.startswith("xl/worksheets/") and n.endswith(".xml")]
            changed = False
            repaired: dict[str, bytes] = {}
            for name in sheet_names:
                raw = source.read(name)
                try:
                    root = ElementTree.fromstring(raw)
                except ElementTree.ParseError:
                    continue
                removed = 0
                for cols in root.iter(f"{{{_MAIN_NS}}}cols"):
                    for col in list(cols):
                        if col.tag != f"{{{_MAIN_NS}}}col":
                            continue
                        try:
                            minimum = int(col.attrib.get("min", "0"))
                            maximum = int(col.attrib.get("max", str(minimum)))
                        except (TypeError, ValueError):
                            cols.remove(col)
                            removed += 1
                            continue
                        if not (1 <= minimum <= maximum <= _XLSX_MAX_COLUMN):
                            cols.remove(col)
                            removed += 1
                if removed:
                    changed = True
                    repaired[name] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=False)
            if not changed:
                return None
            out = io.BytesIO()
            with ZipFile(out, "w", ZIP_DEFLATED) as target:
                for info in source.infolist():
                    target.writestr(info, repaired.get(info.filename, source.read(info.filename)))
            return out.getvalue()
    except (BadZipFile, OSError, ElementTree.ParseError):
        return None


def _load_workbook_resilient(file_bytes: bytes):
    """读取 xlsx；遇到非法列定义时先修复列宽 XML 再重试。"""
    try:
        return load_workbook(io.BytesIO(file_bytes), data_only=True)
    except ValueError:
        repaired = _repair_invalid_column_dimensions(file_bytes)
        if repaired is None:
            raise
        return load_workbook(io.BytesIO(repaired), data_only=True)


def build_template(sheets: list[Sheet]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    hint_fill = PatternFill("solid", fgColor="F5F5F5")
    hint_font = Font(color="888888", size=10)
    for spec in sheets:
        ws = wb.create_sheet(spec.name)
        for idx, col in enumerate(spec.columns, start=1):
            title = f"*{col.label}" if col.required else col.label
            c = ws.cell(row=1, column=idx, value=title)
            c.font = header_font
            hint = col.hint
            if col.enum:
                hint = (hint + "；" if hint else "") + "可选值：" + "/".join(col.enum)
            if col.kind == "date":
                hint = (hint + "；" if hint else "") + "日期格式 2026-01-31"
            h = ws.cell(row=2, column=idx, value=hint or None)
            h.fill = hint_fill
            h.font = hint_font
            ws.column_dimensions[get_column_letter(idx)].width = max(14, len(title) * 2 + 4)
        ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export(spec: Sheet, rows: list[dict]) -> bytes:
    """生成带有说明行的可回导 Excel，导出后可直接修改并再次导入。"""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(spec.name)
    header_font = Font(bold=True)
    hint_fill = PatternFill("solid", fgColor="F5F5F5")
    hint_font = Font(color="888888", size=10)
    for idx, col in enumerate(spec.columns, start=1):
        title = f"*{col.label}" if col.required else col.label
        ws.cell(row=1, column=idx, value=title).font = header_font
        hint = col.hint
        if col.enum:
            hint = (hint + "；" if hint else "") + "可选值：" + "/".join(col.enum)
        h = ws.cell(row=2, column=idx, value=hint or None)
        h.fill = hint_fill
        h.font = hint_font
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(title) * 2 + 4)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, col in enumerate(spec.columns, start=1):
            value = row.get(col.key)
            if isinstance(value, list):
                value = "；".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce(col: Col, value: Any) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if col.kind == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            raise ValueError(f"{col.label}：日期格式应为 2026-01-31")
    if col.kind == "float":
        try:
            return float(value)
        except (TypeError, ValueError):
            raise ValueError(f"{col.label}：应为数字")
    if col.kind == "int":
        try:
            return int(float(value))
        except (TypeError, ValueError):
            raise ValueError(f"{col.label}：应为整数")
    text = str(value).strip()
    if col.enum and text not in col.enum:
        raise ValueError(f"{col.label}：'{text}' 不在可选值 {'/'.join(col.enum)} 中")
    if col.max_length is not None and len(text) > col.max_length:
        raise ValueError(f"{col.label}：不能超过 {col.max_length} 个字符")
    return text


def parse_sheet(file_bytes: bytes, spec: Sheet) -> tuple[list[dict], list[dict]]:
    """返回 (rows, errors)。rows 元素含 _row 行号；说明行(第2行)与空行跳过。"""
    try:
        wb = _load_workbook_resilient(file_bytes)
    except (BadZipFile, InvalidFileException, OSError, KeyError, ValueError):
        # xlsx 本质是 zip；浏览器扩展名校验无法防止损坏文件或把 xls 改名为 xlsx。
        # 将其作为可展示的导入错误返回，避免未处理异常变成前端的 500。
        return [], [{"row": 0, "error": "文件不是有效的 Excel .xlsx 文件，请重新下载模板后填写"}]
    if spec.name not in wb.sheetnames:
        return [], [{"row": 0, "error": f"缺少工作表「{spec.name}」（请使用系统导出的模板）"}]
    ws = wb[spec.name]
    rows: list[dict] = []
    errors: list[dict] = []
    for row_idx, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx == 2:  # 说明行
            continue
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in cells):
            continue
        record: dict = {"_row": row_idx}
        row_errors = []
        for idx, col in enumerate(spec.columns):
            raw = cells[idx] if idx < len(cells) else None
            try:
                value = _coerce(col, raw)
            except ValueError as e:
                row_errors.append(str(e))
                continue
            if col.required and value is None:
                row_errors.append(f"{col.label}：必填")
                continue
            record[col.key] = value
        if row_errors:
            errors.append({"row": row_idx, "error": "；".join(row_errors)})
        else:
            rows.append(record)
    return rows, errors
